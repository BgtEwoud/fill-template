from flask import Flask, render_template, request, redirect, url_for, send_file
import os
from openpyxl import load_workbook
from docxtpl import DocxTemplate
import glob
import pandas as pd

app = Flask(__name__)

# Function to create dictionary from excel file (existing function)
def create_dictionary_from_excel(excel_file_path):
    df = pd.read_excel(excel_file_path, header=None)
    column_values = df.iloc[:, 0].tolist()
    to_fill_in = {value: None for value in column_values}
    return to_fill_in

# Function to generate report (existing function)
def generate_report(docx_file, excel_file, output_folder):
    try:
        to_fill_in = create_dictionary_from_excel(excel_file)
        workbook = load_workbook(excel_file)
        template = DocxTemplate(docx_file)
        worksheet = workbook.active

        for col in range(2, worksheet.max_column + 1):
            for row, key in enumerate(to_fill_in, start=1):
                cell_value = worksheet.cell(row=row, column=col).value
                to_fill_in[key] = cell_value

            template.render(to_fill_in)
            filename = f"{to_fill_in['NAAM_VENNOOTSCHAP']}{os.path.basename(docx_file).replace('Template', '')}.docx"
            filled_path = os.path.join(output_folder, filename)
            template.save(filled_path)
            print(f"Done with {to_fill_in['NAAM_VENNOOTSCHAP']}")
    except Exception as e:
        print(f"Error processing {docx_file}: {e}")

# Route for the homepage (form to upload files)
@app.route('/')
def index():
    return render_template('index.html')

# Route to handle file uploads and processing
@app.route('/generate', methods=['POST'])
def generate():
    if request.method == 'POST':
        input_folder = request.files.getlist('docx_files')  # Get uploaded .docx files
        excel_file = request.files['excel_file']  # Get uploaded Excel file
        output_folder = "output"  # You can set this dynamically or keep it fixed

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Save the uploaded files to a temporary location
        excel_path = os.path.join(output_folder, excel_file.filename)
        excel_file.save(excel_path)

        for docx in input_folder:
            docx_path = os.path.join(output_folder, docx.filename)
            docx.save(docx_path)
            generate_report(docx_path, excel_path, output_folder)  # Call your existing generate_report function

        # After generating, redirect or show a download link
        return redirect(url_for('download', folder=output_folder))

# Route to download the generated reports
@app.route('/download/<folder>')
def download(folder):
    # Serve the generated files for download (for simplicity, one file here)
    files = os.listdir(folder)
    if files:
        file_path = os.path.join(folder, files[0])
        return send_file(file_path, as_attachment=True)
    return "No files generated."

if __name__ == '__main__':
    app.run(debug=True)
